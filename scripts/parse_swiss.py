#!/usr/bin/env python3
import os
import json
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))

SWISS_XLSX = os.path.join(PROJECT_ROOT, 'src', 'data', 'Swiss_food_composition_database.xlsx')
OUTPUT_JSON = os.path.join(PROJECT_ROOT, 'src', 'data', 'swiss_foods.json')

def clean_num(val):
    if val is None or val == '' or val == '-':
        return 0.0
    try:
        return round(float(val), 1)
    except (ValueError, TypeError):
        return 0.0

def parse_swiss():
    print(f"Parsing Swiss Food Composition Database from {SWISS_XLSX}...")
    wb = openpyxl.load_workbook(SWISS_XLSX, data_only=True)
    foods = []

    for sheetname in ['Generic Foods', 'Branded foods']:
        if sheetname not in wb.sheetnames:
            continue
        sheet = wb[sheetname]
        rows = list(sheet.iter_rows(values_only=True))[3:] # skip title and headers

        for r in rows:
            food_id = r[0]
            name = r[3]
            if not food_id or not name:
                continue

            synonyms = r[4] or ''
            category = r[5] or ''
            kcal = clean_num(r[11])
            fat = clean_num(r[14])
            carbs = clean_num(r[41])
            fiber = clean_num(r[50])
            protein = clean_num(r[53])
            sodium = clean_num(r[116])

            # Fallback energy calculation from macros if kcal missing/unmeasured
            if kcal == 0.0 and (protein > 0 or carbs > 0 or fat > 0):
                kcal = round(protein * 4.0 + carbs * 4.0 + fat * 9.0, 1)

            foods.append({
                'id': f'swiss_{food_id}',
                'code': str(food_id),
                'name': str(name).strip(),
                'synonyms': str(synonyms).strip() if synonyms else '',
                'category': str(category).strip() if category else '',
                'servingSizeG': 100,
                'calories100g': kcal,
                'proteins100g': protein,
                'carbs100g': carbs,
                'fats100g': fat,
                'fiber100g': fiber,
                'sodiumMg100g': sodium,
                'source': 'SWISS',
            })

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as out_f:
        json.dump(foods, out_f, ensure_ascii=False, separators=(',', ':'))

    print(f"Successfully generated {OUTPUT_JSON} with {len(foods)} Swiss food items.")

if __name__ == '__main__':
    parse_swiss()
